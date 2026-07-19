"""API 路由 — 工单（公司提单 / 取证调度）"""
from __future__ import annotations

import os
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

import aiofiles
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import case, func, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from config import SCRIPTS_DIR, PROJECT_ROOT
from database import get_db
from models import EvidenceRecord, LinkBatch, Task, VideoLink, WorkOrder, WorkOrderAttachment

router = APIRouter(tags=["工单"])

WORK_ORDER_UPLOAD_DIR = PROJECT_ROOT / "evidence_data" / "work_orders"
WORK_ORDER_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


_SAFE_NAME_RE = re.compile(r'[\\/:*?"<>|]+')
# 主名（别名）/ 主名(别名) → 只保留括号外主名
_PAREN_ALIAS_RE = re.compile(r"[（(][^）)]*[）)]")


def _safe_keyword(name: str) -> str:
    s = (name or "").strip()
    s = _SAFE_NAME_RE.sub("_", s)
    return s[:180]


def _canonical_drama_name(name: str) -> str:
    """只保留主名：去掉中英文括号及其中的别名/营销后缀。"""
    s = (name or "").strip()
    if not s:
        return ""
    s = _PAREN_ALIAS_RE.sub("", s)
    s = re.sub(r"\s+", " ", s).strip(" \t-—–·:：")
    return s


def _decode_text_bytes(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _install_script_text(keyword: str, text: str) -> Path:
    """写入 SCRIPTS_DIR/{keyword}/_script_raw.txt，供 ASR 比对。"""
    kw = _safe_keyword(keyword)
    if not kw:
        raise ValueError("关键词为空")
    script_dir = SCRIPTS_DIR / kw
    script_dir.mkdir(parents=True, exist_ok=True)
    dest = script_dir / "_script_raw.txt"
    body = text.strip()
    dest.write_text(body + ("\n" if body else ""), encoding="utf-8")
    return dest


def _parse_keywords_from_excel(content: bytes) -> list[str]:
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_idx = None
    for i, h in enumerate(header):
        hl = h.lower()
        if any(k in h for k in ("关键词", "剧名", "作品名")) or any(
            k in hl for k in ("keyword", "drama")
        ):
            col_idx = i
            break
    if col_idx is None:
        col_idx = 0

    keywords: list[str] = []
    seen: set[str] = set()
    for row in rows[1:]:
        if not row or col_idx >= len(row):
            continue
        val = row[col_idx]
        if val is None:
            continue
        kw = _canonical_drama_name(str(val))
        if not kw or kw in seen:
            continue
        seen.add(kw)
        keywords.append(kw)
    return keywords


def _index_zip_scripts(zip_bytes: bytes) -> dict[str, bytes]:
    """从 zip 建立 关键词(文件名茎) -> 台词 bytes 索引。"""
    import zipfile
    from io import BytesIO

    index: dict[str, bytes] = {}
    with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            name = info.filename.replace("\\", "/")
            if name.startswith("__MACOSX/") or "/." in f"/{name}":
                continue
            lower = name.lower()
            if not (lower.endswith(".txt") or lower.endswith(".text")):
                continue
            parts = [p for p in name.split("/") if p]
            if not parts:
                continue
            stem = Path(parts[-1]).stem
            if len(parts) >= 2 and stem in ("_script_raw", "script", "台词", "剧本"):
                key = parts[-2]
            else:
                key = stem
            key = key.strip()
            if not key:
                continue
            data = zf.read(info)
            if key in index:
                index[key] = index[key] + b"\n" + data
            else:
                index[key] = data
    return index


def _match_scripts_to_keywords(keywords: list[str], zip_index: dict[str, bytes]) -> dict:
    matched = []
    missing = []
    used_keys: set[str] = set()
    for kw in keywords:
        if kw in zip_index:
            matched.append({"keyword": kw, "zip_key": kw})
            used_keys.add(kw)
            continue
        hit = None
        for zk in zip_index:
            if zk.replace(" ", "") == kw.replace(" ", ""):
                hit = zk
                break
        if hit:
            matched.append({"keyword": kw, "zip_key": hit})
            used_keys.add(hit)
        else:
            missing.append(kw)
    unused = [k for k in zip_index if k not in used_keys]
    return {"matched": matched, "missing": missing, "unused_in_zip": unused}


_URL_RE = re.compile(r"https?://[^\s<>\"{}|\\^`\[\]]+", re.I)

STATUS_LABELS = {
    "draft": "草稿",
    "submitted": "已提交",
    "collecting": "取证中",
    "partial": "部分完成",
    "completed": "已完成",
    "closed": "已关闭",
}


class WorkOrderCreate(BaseModel):
    drama_name: str = Field(..., max_length=200)
    description: str = ""
    priority: int = Field(0, ge=0, le=10)
    deadline: Optional[datetime] = None
    submitter: str = Field("公司用户", max_length=100)


class WorkOrderUpdate(BaseModel):
    drama_name: Optional[str] = Field(None, max_length=200)
    description: Optional[str] = None
    priority: Optional[int] = Field(None, ge=0, le=10)
    deadline: Optional[datetime] = None


class WorkOrderAssign(BaseModel):
    assigned_to: str = Field("", max_length=100)


class WorkOrderLinksBody(BaseModel):
    links: list[str] = Field(default_factory=list, max_length=500)


def _dedupe_links(urls: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for raw in urls:
        u = raw.strip().rstrip(".,;)]")
        if u and u not in seen:
            seen.add(u)
            out.append(u)
    return out


def _parse_links_from_bytes(content: bytes, filename: str) -> list[str]:
    links: list[str] = []
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xls")):
        try:
            import io
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell:
                            links.extend(_URL_RE.findall(str(cell)))
        except Exception:
            pass
    else:
        text = content.decode("utf-8", errors="ignore")
        links.extend(_URL_RE.findall(text))
    return _dedupe_links(links)



def _extract_docx_text(data: bytes) -> str:
    """从 docx 字节提取纯文本（兼容旧调用）。"""
    from weixin.asr.script_cleaner import extract_docx_text

    return extract_docx_text(data)


_OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _cjk_count(s: str) -> int:
    return sum(1 for ch in s if "\u4e00" <= ch <= "\u9fff")


def _fix_zip_name(name: str) -> str:
    """尽力把 zip 内文件名修成可读中文。"""
    name = (name or "").replace("\\", "/")
    if _cjk_count(name) > 0:
        return name
    # 旧工具：原始字节按 cp437 解出来，再按 gbk/gb18030 还原
    for enc in ("gbk", "gb18030", "utf-8"):
        try:
            fixed = name.encode("cp437", errors="strict").decode(enc)
            if _cjk_count(fixed) > _cjk_count(name):
                return fixed.replace("\\", "/")
        except Exception:
            continue
    # 宽松：忽略无法映射的码点
    try:
        raw = name.encode("cp437", errors="replace")
        # replace 会插入 ?，改用 latin-1 保字节
        raw = name.encode("latin-1", errors="ignore")
        for enc in ("gbk", "gb18030", "utf-8"):
            try:
                fixed = raw.decode(enc)
                if _cjk_count(fixed) > 0:
                    return fixed.replace("\\", "/")
            except Exception:
                continue
    except Exception:
        pass
    return name


def _norm_drama_key(s: str) -> str:
    """剧名/文件名归一化，便于模糊匹配。"""
    t = (s or "").strip().lower()
    for ch in (
        " ", "\u3000", "_", "-", "—", "–", "·", ".", "。", "，", ",",
        "（", "）", "(", ")", "【", "】", "[", "]", "《", "》", ":", "：",
        "?", "？", "!", "！",
    ):
        t = t.replace(ch, "")
    return t


def _detect_script(filename: str, data: bytes) -> str | None:
    """返回脚本扩展名 .doc/.docx，非脚本则 None。"""
    lower = (filename or "").lower().replace("\\", "/")
    base = lower.rsplit("/", 1)[-1]
    if base.startswith("~"):
        return None
    if lower.endswith(".docx"):
        return ".docx"
    if lower.endswith(".doc"):
        return ".doc"
    # 文件名乱码丢了后缀时，靠魔数识别
    if data[:8] == _OLE_MAGIC:
        return ".doc"
    if data[:2] == b"PK" and (b"word/" in data[:4096] or b"WordDocument" in data[:4096]):
        return ".docx"
    return None


def _scan_zip_once(zip_bytes: bytes, metadata_encoding: str | None) -> tuple[list[tuple[str, bytes]], dict[str, tuple[bytes, str]], int]:
    """按指定 metadata_encoding 扫一遍，返回 (excels, scripts, cjk_score)。"""
    import zipfile
    from io import BytesIO

    excels: list[tuple[str, bytes]] = []
    script_index: dict[str, tuple[bytes, str]] = {}
    score = 0
    kwargs = {}
    if metadata_encoding:
        kwargs["metadata_encoding"] = metadata_encoding
    with zipfile.ZipFile(BytesIO(zip_bytes), **kwargs) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            # metadata_encoding 已按 gbk 解时直接用；否则再尝试修复
            name = (info.filename or "").replace("\\", "/")
            if not metadata_encoding:
                name = _fix_zip_name(name)
            if name.startswith("__MACOSX/") or "/." in f"/{name}":
                continue
            score += _cjk_count(name)
            parts = [p for p in name.split("/") if p]
            if not parts:
                continue
            data = zf.read(info)
            lower = name.lower()
            if lower.endswith((".xlsx", ".xls")) or (
                data[:2] == b"PK" and b"xl/" in data[:4096]
            ):
                # 表格以扩展名优先；魔数兜底时用原名
                excel_name = parts[-1] if lower.endswith((".xlsx", ".xls")) else (parts[-1] + ".xlsx")
                excels.append((excel_name, data))
                continue
            ext = _detect_script(name, data)
            if not ext:
                continue
            stem = Path(parts[-1]).stem.strip().rstrip(".?")
            if not stem or _cjk_count(stem) == 0:
                # 文件名仍乱码：用占位 stem，后续按「唯一剧本」强制配对
                stem = f"script_{len(script_index)+1}"
            if stem in script_index and ext == ".doc":
                continue
            script_index[stem] = (data, ext)
    return excels, script_index, score


def _scan_zip_entries(zip_bytes: bytes) -> tuple[list[tuple[str, bytes]], dict[str, tuple[bytes, str]]]:
    """返回 (excel列表[(name,bytes)], 剧本索引{stem: (bytes, ext)})。

    WinRAR 中文 zip 常用 GBK 文件名；依次尝试 gbk / gb18030 / 默认+修复，取中文最多的结果。
    """
    candidates: list[tuple[int, list, dict]] = []
    for enc in ("gbk", "gb18030", None):
        try:
            excels, scripts, score = _scan_zip_once(zip_bytes, enc)
            # 有剧本/表格时加分，避免空扫赢
            bonus = 10 * len(scripts) + 5 * len(excels)
            candidates.append((score + bonus, excels, scripts))
        except Exception as e:
            print(f"[import-package] zip scan enc={enc!r} fail: {e}")
    if not candidates:
        return [], {}
    candidates.sort(key=lambda x: x[0], reverse=True)
    _, excels, scripts = candidates[0]
    return excels, scripts


def _match_docx_to_keywords(keywords: list[str], script_index: dict) -> dict:
    matched = []
    missing = []
    used: set[str] = set()

    def _find_hit(kw: str) -> str | None:
        if kw in script_index and kw not in used:
            return kw
        nkw = _norm_drama_key(kw)
        for zk in script_index:
            if zk in used:
                continue
            if _norm_drama_key(zk) == nkw:
                return zk
        best = None
        best_score = 0
        for zk in script_index:
            if zk in used:
                continue
            nzk = _norm_drama_key(zk)
            if not nzk or not nkw:
                continue
            if nkw in nzk or nzk in nkw:
                score = min(len(nkw), len(nzk))
                if score >= 2 and score > best_score:
                    best, best_score = zk, score
        return best

    for kw in keywords:
        hit = _find_hit(kw)
        if hit:
            matched.append({"keyword": kw, "docx_key": hit})
            used.add(hit)
        else:
            missing.append(kw)

    # 强制兜底：剩余剧名与剩余剧本按体积从大到小配对（彻底无视文件名乱码）
    unused = [k for k in script_index if k not in used]
    unused.sort(key=lambda k: len(script_index[k][0]), reverse=True)
    while missing and unused:
        kw = missing.pop(0)
        hit = unused.pop(0)
        matched.append({"keyword": kw, "docx_key": hit})
        used.add(hit)

    return {
        "matched": matched,
        "missing": missing,
        "unused_docx": [k for k in script_index if k not in used],
    }


def _parse_help_collect_excel(content: bytes) -> list[dict]:
    """解析帮我取证 Excel，返回 [{platform, drama_name, link}, ...]。"""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    platform_idx = None
    drama_idx = None
    link_idx = None
    for i, h in enumerate(header):
        hl = h.lower()
        if platform_idx is None and (
            "平台" in h or "platform" in hl
        ):
            platform_idx = i
        if drama_idx is None and (
            any(k in h for k in ("剧名", "关键词", "作品名"))
            or any(k in hl for k in ("drama", "keyword"))
        ):
            drama_idx = i
        if link_idx is None and (
            any(k in h for k in ("侵权视频链接", "视频链接", "链接"))
            or "link" in hl
            or "url" in hl
        ):
            link_idx = i

    if platform_idx is None or drama_idx is None or link_idx is None:
        raise ValueError(
            f"Excel 需同时包含「平台」「剧名」「侵权视频链接」列，实际表头: {header}"
        )

    items: list[dict] = []
    for row in rows[1:]:
        if not row:
            continue
        platform = ""
        drama = ""
        link = ""
        if platform_idx < len(row) and row[platform_idx] is not None:
            platform = str(row[platform_idx]).strip()
        if drama_idx < len(row) and row[drama_idx] is not None:
            drama = _canonical_drama_name(str(row[drama_idx]))
        if link_idx < len(row) and row[link_idx] is not None:
            link = str(row[link_idx]).strip()
        if not platform and not drama and not link:
            continue
        urls = _URL_RE.findall(link) if link else []
        url = urls[0] if urls else link
        items.append({"platform": platform, "drama_name": drama, "link": url})
    return items


# 一期开放的平台（中文名 / 别名 → 标准名）
ALLOWED_PLATFORM_ALIASES = {
    "微信视频号": "微信视频号",
    "视频号": "微信视频号",
    "微信": "微信视频号",
    "weixin": "微信视频号",
    "weixin_channels": "微信视频号",
    "wechat": "微信视频号",
}


def _normalize_platform(raw: str) -> str | None:
    """返回标准平台名；未识别返回 None。"""
    s = (raw or "").strip()
    if not s:
        return None
    if s in ALLOWED_PLATFORM_ALIASES:
        return ALLOWED_PLATFORM_ALIASES[s]
    low = s.lower()
    if low in ALLOWED_PLATFORM_ALIASES:
        return ALLOWED_PLATFORM_ALIASES[low]
    return None


ALLOWED_PLATFORMS = {"weixin_channels": "微信视频号"}


async def _get_or_create_wo_batch(db: AsyncSession, wo: WorkOrder) -> LinkBatch:
    batch_name = f"WO-{wo.order_no}"
    batch = (
        await db.execute(select(LinkBatch).where(LinkBatch.name == batch_name))
    ).scalar_one_or_none()
    if not batch:
        batch = LinkBatch(name=batch_name, source="work_order")
        db.add(batch)
        await db.flush()
    return batch


async def _import_links_to_batch(
    db: AsyncSession,
    wo: WorkOrder,
    links: list[str],
) -> dict:
    """将链接写入工单专属链接池批次（跳过已存在 URL）"""
    links = _dedupe_links(links)
    if not links:
        return {"imported": 0, "skipped": 0, "batch_id": None, "batch_name": ""}

    batch = await _get_or_create_wo_batch(db, wo)
    existing = set(
        (await db.execute(
            select(VideoLink.link_url).where(VideoLink.batch_id == batch.id)
        )).scalars().all()
    )

    imported = 0
    skipped = 0
    max_order = (
        await db.execute(
            select(func.max(VideoLink.sort_order)).where(VideoLink.batch_id == batch.id)
        )
    ).scalar() or 0

    for url in links:
        if url in existing:
            skipped += 1
            continue
        max_order += 1
        db.add(VideoLink(
            batch_id=batch.id,
            keyword=wo.drama_name,
            link_url=url,
            sort_order=max_order,
        ))
        existing.add(url)
        imported += 1

    total = (
        await db.execute(
            select(func.count(VideoLink.id)).where(VideoLink.batch_id == batch.id)
        )
    ).scalar() or 0
    batch.total_count = total
    wo.updated_at = datetime.now()
    await db.flush()
    return {
        "imported": imported,
        "skipped": skipped,
        "batch_id": batch.id,
        "batch_name": batch.name,
        "total_in_batch": total,
    }


async def _batch_info_for_work_order(db: AsyncSession, wo: WorkOrder) -> dict:
    batch_name = f"WO-{wo.order_no}"
    batch = (
        await db.execute(select(LinkBatch).where(LinkBatch.name == batch_name))
    ).scalar_one_or_none()
    if not batch:
        return {"batch_id": None, "batch_name": batch_name, "link_count": 0, "pending_count": 0}
    pending = (
        await db.execute(
            select(func.count(VideoLink.id)).where(
                VideoLink.batch_id == batch.id,
                VideoLink.evidence_record_id.is_(None),
            )
        )
    ).scalar() or 0
    total = (
        await db.execute(
            select(func.count(VideoLink.id)).where(VideoLink.batch_id == batch.id)
        )
    ).scalar() or 0
    return {
        "batch_id": batch.id,
        "batch_name": batch.name,
        "link_count": total,
        "pending_count": pending,
    }


def _gen_order_no() -> str:
    return f"WO{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4].upper()}"


async def _refresh_work_order_stats(db: AsyncSession, work_order_id: int) -> None:
    """根据关联任务刷新工单证据数 / 推送数 / 状态。"""
    wo = (await db.execute(select(WorkOrder).where(WorkOrder.id == work_order_id))).scalar_one_or_none()
    if not wo:
        return

    task_ids = (
        await db.execute(select(Task.id).where(Task.work_order_id == work_order_id))
    ).scalars().all()
    if not task_ids:
        return

    ev_q = select(
        func.count(EvidenceRecord.id),
        func.sum(case((EvidenceRecord.pushed_to_company == True, 1), else_=0)),
        func.sum(case((EvidenceRecord.pushed_to_police == True, 1), else_=0)),
    ).where(EvidenceRecord.task_id.in_(task_ids))
    row = (await db.execute(ev_q)).first()
    evidence_count = int(row[0] or 0) if row else 0
    company_pushed_count = int(row[1] or 0) if row else 0
    pushed_count = int(row[2] or 0) if row else 0

    wo.evidence_count = evidence_count
    wo.company_pushed_count = company_pushed_count
    wo.pushed_count = pushed_count

    running = (
        await db.execute(
            select(func.count(Task.id)).where(
                Task.work_order_id == work_order_id,
                Task.status.in_(["running", "links_collected", "pending"]),
            )
        )
    ).scalar() or 0

    if wo.status in ("submitted", "collecting", "partial"):
        if running > 0:
            wo.status = "collecting"
            if not wo.started_at:
                wo.started_at = datetime.now()
        elif evidence_count > 0:
            wo.status = "partial"
        elif wo.status == "collecting" and evidence_count == 0 and running == 0:
            wo.status = "submitted"

    await db.flush()


def _wo_to_dict(wo: WorkOrder, attachments: list | None = None) -> dict:
    from engine.script_clean_jobs import SCRIPT_STATUS_LABELS

    script_status = getattr(wo, "script_status", None) or "none"
    return {
        "id": wo.id,
        "order_no": wo.order_no,
        "drama_name": wo.drama_name,
        "description": wo.description or "",
        "priority": wo.priority or 0,
        "deadline": str(wo.deadline) if wo.deadline else "",
        "status": wo.status,
        "status_label": STATUS_LABELS.get(wo.status, wo.status),
        "submitter": wo.submitter or "",
        "assigned_to": wo.assigned_to or "",
        "evidence_count": wo.evidence_count or 0,
        "company_pushed_count": getattr(wo, "company_pushed_count", None) or 0,
        "pushed_count": wo.pushed_count or 0,
        "script_status": script_status,
        "script_status_label": SCRIPT_STATUS_LABELS.get(script_status, script_status),
        "script_source_hash": getattr(wo, "script_source_hash", None) or "",
        "script_cleaned_at": str(wo.script_cleaned_at) if getattr(wo, "script_cleaned_at", None) else "",
        "script_error": getattr(wo, "script_error", None) or "",
        "submitted_at": str(wo.submitted_at) if wo.submitted_at else "",
        "started_at": str(wo.started_at) if wo.started_at else "",
        "completed_at": str(wo.completed_at) if wo.completed_at else "",
        "created_at": str(wo.created_at) if wo.created_at else "",
        "updated_at": str(wo.updated_at) if wo.updated_at else "",
        "attachments": attachments or [],
    }


@router.get("/work-orders")
async def list_work_orders(
    status: str = Query("", max_length=20),
    queue: str = Query("", description="pending|mine|all"),
    assigned_to: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    base = select(WorkOrder)
    if status:
        base = base.where(WorkOrder.status == status)
    if queue == "pending":
        base = base.where(WorkOrder.status.in_(["submitted", "collecting", "partial"]))
    elif queue == "mine" and assigned_to:
        base = base.where(
            WorkOrder.assigned_to == assigned_to,
            WorkOrder.status.in_(["submitted", "collecting", "partial"]),
        )
    base = base.order_by(WorkOrder.priority.desc(), WorkOrder.submitted_at.desc(), WorkOrder.id.desc())

    count_q = select(func.count(WorkOrder.id))
    if status:
        count_q = count_q.where(WorkOrder.status == status)
    if queue == "pending":
        count_q = count_q.where(WorkOrder.status.in_(["submitted", "collecting", "partial"]))
    elif queue == "mine" and assigned_to:
        count_q = count_q.where(
            WorkOrder.assigned_to == assigned_to,
            WorkOrder.status.in_(["submitted", "collecting", "partial"]),
        )
    total = (await db.execute(count_q)).scalar() or 0
    rows = (await db.execute(base.offset((page - 1) * page_size).limit(page_size))).scalars().all()
    items = []
    for r in rows:
        d = _wo_to_dict(r)
        d["link_pool"] = await _batch_info_for_work_order(db, r)
        items.append(d)
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("/work-orders", status_code=201)
async def create_work_order(body: WorkOrderCreate, db: AsyncSession = Depends(get_db)):
    drama = _canonical_drama_name(body.drama_name)
    if not drama:
        raise HTTPException(400, "剧名不能为空")
    wo = WorkOrder(
        order_no=_gen_order_no(),
        drama_name=drama,
        description=body.description or "",
        priority=body.priority,
        deadline=body.deadline,
        status="draft",
        submitter=body.submitter,
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )
    db.add(wo)
    await db.commit()
    await db.refresh(wo)
    return _wo_to_dict(wo)


@router.get("/work-orders/package-template")
async def download_package_template():
    """下载新建工单样板 zip（剧名 Excel + 示例 docx）。"""
    import zipfile
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from docx import Document

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "剧名"
        ws.append(["剧名"])
        ws.append(["示例剧名A"])
        ws.append(["示例剧名B"])
        xbuf = BytesIO()
        wb.save(xbuf)
        zf.writestr("工单包/剧名列表.xlsx", xbuf.getvalue())

        # Sample docx A
        for name, sample in (
            ("示例剧名A", "这是示例剧名A的台词第一句。\n这是第二句对白。"),
            ("示例剧名B", "这是示例剧名B的台词第一句。\n请替换为真实剧本。"),
        ):
            doc = Document()
            for line in sample.split("\n"):
                doc.add_paragraph(line)
            dbuf = BytesIO()
            doc.save(dbuf)
            zf.writestr(f"工单包/{name}.docx", dbuf.getvalue())

        tip = (
            "打包说明\n"
            "1. 新建文件夹（如「工单包」）\n"
            "2. 放入 Excel：每行一部剧名，列名为「剧名」\n"
            "3. 同目录放置台词 docx：一部剧一个文件，文件名=剧名\n"
            "4. 将整个文件夹打成 zip 后，在「新建工单」页上传\n"
        )
        zf.writestr("工单包/README.txt", tip.encode("utf-8"))

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={
            "Content-Disposition": "attachment; filename=work-order-package-sample.zip"
        },
    )


@router.post("/work-orders/import-package")
async def import_work_order_package(
    file: UploadFile = File(..., description="工单包 zip"),
    submitter: str = Form("公司用户"),
    db: AsyncSession = Depends(get_db),
):
    """新建工单：上传 zip 入库（不清洗台词）。缺台词也建单，仅告知；清洗延后到取证认领。"""
    from engine.script_clean_jobs import (
        SCRIPT_STATUS_NONE,
        SCRIPT_STATUS_PENDING,
        SCRIPT_STATUS_READY,
        library_ready_for_hash,
        sha256_bytes,
    )

    fname = Path(file.filename or "package.zip").name
    if not fname.lower().endswith(".zip"):
        raise HTTPException(400, "请上传 .zip 工单包")

    zip_bytes = await file.read()
    try:
        excels, script_index = _scan_zip_entries(zip_bytes)
    except Exception as e:
        raise HTTPException(400, f"无法读取 zip: {e}") from e

    if not excels:
        raise HTTPException(400, "zip 中未找到 Excel（.xlsx），请放入剧名列表")
    excel_bytes = None
    excel_name = ""
    for n, data in excels:
        if any(k in n for k in ("剧名", "关键词", "drama", "keyword")):
            excel_bytes, excel_name = data, n
            break
    if excel_bytes is None:
        excel_name, excel_bytes = excels[0]

    keywords = _parse_keywords_from_excel(excel_bytes)
    if not keywords:
        raise HTTPException(400, f"「{excel_name}」中未解析到剧名")

    pairing = _match_docx_to_keywords(keywords, script_index)
    matched_map = {m["keyword"]: m for m in pairing["matched"]}
    missing_script = list(pairing["missing"])
    script_pending: list[str] = []
    script_ready: list[str] = []
    created = []
    now = datetime.now()
    import_id = uuid.uuid4().hex[:12]

    # 整包留存
    import_dir = WORK_ORDER_UPLOAD_DIR / "imports" / import_id
    import_dir.mkdir(parents=True, exist_ok=True)
    package_path = import_dir / "package.zip"
    package_path.write_bytes(zip_bytes)

    print(
        f"[import-package] import_id={import_id} excel={excel_name!r} "
        f"keywords={keywords} scripts={list(script_index.keys())} "
        f"matched={pairing['matched']} missing={pairing['missing']}"
    )

    for kw in keywords:
        hit = matched_map.get(kw)
        script_status = SCRIPT_STATUS_NONE
        source_hash = ""
        att_info = None

        if hit:
            raw, ext = script_index[hit["docx_key"]]
            source_hash = sha256_bytes(raw)
            if library_ready_for_hash(kw, source_hash):
                script_status = SCRIPT_STATUS_READY
                script_ready.append(kw)
            else:
                script_status = SCRIPT_STATUS_PENDING
                script_pending.append(kw)
            att_info = (raw, ext, f"{hit['docx_key']}{ext}", hit["docx_key"])

        wo = WorkOrder(
            order_no=_gen_order_no(),
            drama_name=kw,
            description=f"由工单包导入（{fname}，import={import_id}）",
            status="submitted",
            priority=5,
            submitter=submitter,
            submitted_at=now,
            created_at=now,
            updated_at=now,
            script_status=script_status,
            script_source_hash=source_hash,
            script_cleaned_at=now if script_status == SCRIPT_STATUS_READY else None,
            script_error="",
        )
        db.add(wo)
        await db.flush()

        if att_info:
            raw, ext, script_name, stem = att_info
            dest_dir = WORK_ORDER_UPLOAD_DIR / str(wo.id)
            dest_dir.mkdir(parents=True, exist_ok=True)
            safe_stem = _safe_keyword(stem) or "script"
            att_path = dest_dir / f"{uuid.uuid4().hex[:8]}_{safe_stem}{ext}"
            att_path.write_bytes(raw)
            db.add(WorkOrderAttachment(
                work_order_id=wo.id,
                file_name=script_name,
                file_type="script",
                file_path=str(att_path.relative_to(PROJECT_ROOT)).replace("\\", "/"),
                file_size=len(raw),
                created_at=now,
            ))

        created.append({
            "id": wo.id,
            "order_no": wo.order_no,
            "drama_name": kw,
            "script_status": script_status,
            "has_script_file": bool(att_info),
        })

    await db.commit()

    unused = pairing["unused_docx"]
    parts = [f"已创建 {len(created)} 张工单"]
    if script_pending:
        parts.append(f"{len(script_pending)} 部台词待取证端认领后清洗")
    if script_ready:
        parts.append(f"{len(script_ready)} 部台词已就绪可复用")
    if missing_script:
        parts.append(f"{len(missing_script)} 部缺台词文件（仍已建单）")
    if unused:
        parts.append(f"未匹配剧本: {', '.join(unused)}")

    return {
        "created": created,
        "created_count": len(created),
        "missing_script": missing_script,
        "script_pending": script_pending,
        "script_ready": script_ready,
        "unused_docx": unused,
        "script_files": list(script_index.keys()),
        "excel_name": excel_name,
        "import_id": import_id,
        "package_path": str(package_path.relative_to(PROJECT_ROOT)).replace("\\", "/"),
        "message": "；".join(parts),
    }


@router.post("/work-orders/help-collect")
async def help_collect(
    excel: UploadFile = File(..., description="侵权链接 Excel"),
    submitter: str = Form("公司用户"),
    db: AsyncSession = Depends(get_db),
):
    """帮我取证：上传含「平台、剧名、侵权视频链接」的 Excel，按剧名挂到已有工单链接池。"""
    fname = Path(excel.filename or "links.xlsx").name
    if not fname.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传 .xlsx / .xls")

    content = await excel.read()
    try:
        rows = _parse_help_collect_excel(content)
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    if not rows:
        raise HTTPException(400, "Excel 中没有有效数据行")

    # 剧名 -> 最新工单
    drama_names = {r["drama_name"] for r in rows if r["drama_name"]}
    wo_map: dict[str, WorkOrder] = {}
    if drama_names:
        result = await db.execute(
            select(WorkOrder)
            .where(WorkOrder.drama_name.in_(list(drama_names)))
            .order_by(WorkOrder.id.desc())
        )
        for wo in result.scalars().all():
            if wo.drama_name not in wo_map:
                wo_map[wo.drama_name] = wo

    success = []
    failed = []
    platforms_used: set[str] = set()

    # 按工单聚合链接，并记录平台
    by_wo: dict[int, list[str]] = {}
    wo_obj: dict[int, WorkOrder] = {}
    wo_platforms: dict[int, set[str]] = {}
    for r in rows:
        drama = r["drama_name"]
        link = r["link"]
        plat_raw = r.get("platform") or ""
        plat = _normalize_platform(plat_raw)
        if not plat_raw.strip():
            failed.append({"drama_name": drama, "link": link, "reason": "缺少平台"})
            continue
        if plat is None:
            failed.append({
                "drama_name": drama,
                "link": link,
                "reason": f"暂不支持的平台「{plat_raw}」（一期仅微信视频号）",
            })
            continue
        if not drama:
            failed.append({"drama_name": "", "link": link, "reason": "缺少剧名"})
            continue
        if not link or not link.startswith("http"):
            failed.append({"drama_name": drama, "link": link, "reason": "无效链接"})
            continue
        wo = wo_map.get(drama)
        if not wo:
            wo = next(
                (v for k, v in wo_map.items() if k.replace(" ", "") == drama.replace(" ", "")),
                None,
            )
        if not wo:
            failed.append({"drama_name": drama, "link": link, "reason": "未找到对应工单，请先新建工单"})
            continue
        by_wo.setdefault(wo.id, []).append(link)
        wo_obj[wo.id] = wo
        wo_platforms.setdefault(wo.id, set()).add(plat)
        platforms_used.add(plat)

    for woid, links in by_wo.items():
        wo = wo_obj[woid]
        summary = await _import_links_to_batch(db, wo, links)
        plat_note = "、".join(sorted(wo_platforms.get(woid, set())))
        note = f"[帮我取证] 平台：{plat_note}；导入人：{submitter}"
        desc = wo.description or ""
        if note not in desc:
            wo.description = (desc + "\n" + note).strip() if desc else note
        if wo.status == "submitted":
            wo.status = "collecting"
            if not wo.started_at:
                wo.started_at = datetime.now()
        wo.updated_at = datetime.now()
        success.append({
            "work_order_id": wo.id,
            "order_no": wo.order_no,
            "drama_name": wo.drama_name,
            "platforms": sorted(wo_platforms.get(woid, set())),
            "imported": summary["imported"],
            "skipped": summary["skipped"],
            "batch_name": summary["batch_name"],
        })

    await db.commit()
    imported_total = sum(s["imported"] for s in success)
    return {
        "platforms": sorted(platforms_used),
        "success": success,
        "failed": failed,
        "imported_total": imported_total,
        "failed_count": len(failed),
        "message": (
            f"已导入 {imported_total} 条链接到 {len(success)} 个工单"
            + (f"，{len(failed)} 条失败" if failed else "")
        ),
    }



@router.get("/work-orders/{work_order_id}")
async def get_work_order(work_order_id: int, db: AsyncSession = Depends(get_db)):
    wo = (await db.execute(select(WorkOrder).where(WorkOrder.id == work_order_id))).scalar_one_or_none()
    if not wo:
        raise HTTPException(404, "工单不存在")
    await _refresh_work_order_stats(db, work_order_id)
    await db.commit()
    await db.refresh(wo)

    atts = (
        await db.execute(
            select(WorkOrderAttachment).where(WorkOrderAttachment.work_order_id == work_order_id)
        )
    ).scalars().all()
    att_list = [
        {
            "id": a.id,
            "file_name": a.file_name,
            "file_type": a.file_type,
            "file_path": a.file_path,
            "file_size": a.file_size,
            "created_at": str(a.created_at) if a.created_at else "",
        }
        for a in atts
    ]

    tasks = (
        await db.execute(
            select(Task).where(Task.work_order_id == work_order_id).order_by(Task.id.desc())
        )
    ).scalars().all()
    task_list = [
        {
            "id": t.id,
            "keyword": t.keyword,
            "status": t.status,
            "phase": t.phase,
            "created_at": str(t.created_at) if t.created_at else "",
        }
        for t in tasks
    ]

    # 侵权等级分布（公司端仅统计）
    level_stats = {"high": 0, "mid": 0, "low": 0, "infringement": 0}
    ev_rows: list = []
    if tasks:
        task_ids = [t.id for t in tasks]
        ev_rows = (
            await db.execute(
                select(EvidenceRecord.infringement_level, EvidenceRecord.review_status).where(
                    EvidenceRecord.task_id.in_(task_ids)
                )
            )
        ).all()
        for level, review in ev_rows:
            if review == "侵权":
                level_stats["infringement"] += 1
            if level in ("高度疑似", "侵权"):
                level_stats["high"] += 1
            elif level == "疑似":
                level_stats["mid"] += 1
            elif level == "待观察":
                level_stats["low"] += 1

    data = _wo_to_dict(wo, att_list)
    data["tasks"] = task_list
    data["level_stats"] = level_stats
    data["link_pool"] = await _batch_info_for_work_order(db, wo)
    data["review_stats"] = {
        "infringement": sum(1 for _, review in ev_rows if review == "侵权"),
        "not_infringement": sum(1 for _, review in ev_rows if review == "未侵权"),
        "pending": sum(1 for _, review in ev_rows if not review),
    }
    return data


@router.patch("/work-orders/{work_order_id}")
async def update_work_order(
    work_order_id: int,
    body: WorkOrderUpdate,
    db: AsyncSession = Depends(get_db),
):
    wo = (await db.execute(select(WorkOrder).where(WorkOrder.id == work_order_id))).scalar_one_or_none()
    if not wo:
        raise HTTPException(404, "工单不存在")
    if wo.status not in ("draft", "submitted"):
        raise HTTPException(400, "仅草稿或已提交状态可编辑基本信息")
    if body.drama_name is not None:
        drama = _canonical_drama_name(body.drama_name)
        if not drama:
            raise HTTPException(400, "剧名不能为空")
        wo.drama_name = drama
    if body.description is not None:
        wo.description = body.description
    if body.priority is not None:
        wo.priority = body.priority
    if body.deadline is not None:
        wo.deadline = body.deadline
    wo.updated_at = datetime.now()
    await db.commit()
    await db.refresh(wo)
    return _wo_to_dict(wo)


@router.post("/work-orders/{work_order_id}/submit")
async def submit_work_order(work_order_id: int, db: AsyncSession = Depends(get_db)):
    wo = (await db.execute(select(WorkOrder).where(WorkOrder.id == work_order_id))).scalar_one_or_none()
    if not wo:
        raise HTTPException(404, "工单不存在")
    if wo.status != "draft":
        raise HTTPException(400, "仅草稿可提交")
    wo.status = "submitted"
    wo.submitted_at = datetime.now()
    wo.updated_at = datetime.now()
    atts = (
        await db.execute(
            select(WorkOrderAttachment).where(WorkOrderAttachment.work_order_id == work_order_id)
        )
    ).scalars().all()
    link_atts = [a for a in atts if a.file_type == "links"]
    import_summary = {"imported": 0, "skipped": 0}
    for att in link_atts:
        try:
            full = PROJECT_ROOT / att.file_path
            if full.is_file():
                content = full.read_bytes()
                urls = _parse_links_from_bytes(content, att.file_name)
                r = await _import_links_to_batch(db, wo, urls)
                import_summary["imported"] += r["imported"]
                import_summary["skipped"] += r["skipped"]
        except Exception:
            pass
    await db.commit()
    await db.refresh(wo)
    data = _wo_to_dict(wo)
    data["link_import"] = import_summary
    return data


@router.post("/work-orders/{work_order_id}/assign")
async def assign_work_order(
    work_order_id: int,
    body: WorkOrderAssign,
    db: AsyncSession = Depends(get_db),
):
    from engine.script_clean_jobs import (
        SCRIPT_STATUS_PENDING,
        SCRIPT_STATUS_READY,
        library_ready_for_hash,
        schedule_clean_work_order,
    )

    wo = (await db.execute(select(WorkOrder).where(WorkOrder.id == work_order_id))).scalar_one_or_none()
    if not wo:
        raise HTTPException(404, "工单不存在")
    if wo.status not in ("submitted", "collecting", "partial"):
        raise HTTPException(400, "当前状态不可认领")
    wo.assigned_to = body.assigned_to or "取证员"
    if wo.status == "submitted":
        wo.status = "collecting"
        wo.started_at = datetime.now()
    wo.updated_at = datetime.now()

    # 认领时：有源文件且未就绪 → 触发异步清洗；哈希已命中则直接 ready
    script_status = getattr(wo, "script_status", None) or "none"
    source_hash = getattr(wo, "script_source_hash", None) or ""
    if script_status not in ("none",) and wo.drama_name:
        if source_hash and library_ready_for_hash(wo.drama_name, source_hash):
            wo.script_status = SCRIPT_STATUS_READY
            if not wo.script_cleaned_at:
                wo.script_cleaned_at = datetime.now()
            wo.script_error = ""
        elif script_status in (SCRIPT_STATUS_PENDING, "failed", "cleaning"):
            # cleaning 卡住时也允许重新调度
            if script_status != "cleaning":
                wo.script_status = SCRIPT_STATUS_PENDING

    await db.commit()
    await db.refresh(wo)

    if (getattr(wo, "script_status", None) or "") in (SCRIPT_STATUS_PENDING, "cleaning"):
        schedule_clean_work_order(wo.id)

    return _wo_to_dict(wo)


@router.post("/work-orders/{work_order_id}/attachments")
async def upload_attachment(
    work_order_id: int,
    file_type: str = Form("other"),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    from engine.script_clean_jobs import (
        SCRIPT_STATUS_PENDING,
        SCRIPT_STATUS_READY,
        library_ready_for_hash,
        schedule_clean_work_order,
        sha256_bytes,
    )

    wo = (await db.execute(select(WorkOrder).where(WorkOrder.id == work_order_id))).scalar_one_or_none()
    if not wo:
        raise HTTPException(404, "工单不存在")

    safe_name = Path(file.filename or "upload").name
    dest_dir = WORK_ORDER_UPLOAD_DIR / str(work_order_id)
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_path = dest_dir / f"{uuid.uuid4().hex[:8]}_{safe_name}"

    content = await file.read()
    async with aiofiles.open(dest_path, "wb") as fh:
        await fh.write(content)

    # 剧本：只存附件与哈希，清洗延后（认领或此处触发）
    if file_type == "script" and wo.drama_name:
        lower = safe_name.lower()
        if lower.endswith((".doc", ".docx")):
            source_hash = sha256_bytes(content)
            wo.script_source_hash = source_hash
            if library_ready_for_hash(wo.drama_name, source_hash):
                wo.script_status = SCRIPT_STATUS_READY
                wo.script_cleaned_at = datetime.now()
                wo.script_error = ""
            else:
                wo.script_status = SCRIPT_STATUS_PENDING
                wo.script_cleaned_at = None
                wo.script_error = ""
        elif lower.endswith(".txt") or lower.endswith(".zip"):
            # 兼容直接传纯文本 / txt：仍即时写入 raw（已是对白）
            script_text = ""
            if lower.endswith(".zip"):
                zidx = _index_zip_scripts(content)
                chunks = [_decode_text_bytes(v) for v in zidx.values()]
                script_text = "\n\n".join(chunks)
            else:
                script_text = _decode_text_bytes(content)
            _install_script_text(wo.drama_name, script_text)
            wo.script_status = SCRIPT_STATUS_READY
            wo.script_source_hash = sha256_bytes(content)
            wo.script_cleaned_at = datetime.now()
            wo.script_error = ""

    att = WorkOrderAttachment(
        work_order_id=work_order_id,
        file_name=safe_name,
        file_type=file_type,
        file_path=str(dest_path.relative_to(PROJECT_ROOT)).replace("\\", "/"),
        file_size=len(content),
        created_at=datetime.now(),
    )
    db.add(att)
    wo.updated_at = datetime.now()
    await db.flush()

    link_import = None
    if file_type == "links":
        urls = _parse_links_from_bytes(content, safe_name)
        link_import = await _import_links_to_batch(db, wo, urls)

    await db.commit()
    await db.refresh(wo)

    if file_type == "script" and (getattr(wo, "script_status", None) or "") == SCRIPT_STATUS_PENDING:
        schedule_clean_work_order(wo.id)

    atts = (
        await db.execute(
            select(WorkOrderAttachment).where(WorkOrderAttachment.work_order_id == work_order_id)
        )
    ).scalars().all()
    data = _wo_to_dict(wo, [
        {
            "id": a.id,
            "file_name": a.file_name,
            "file_type": a.file_type,
            "file_path": a.file_path,
            "file_size": a.file_size,
        }
        for a in atts
    ])
    if link_import is not None:
        data["link_import"] = link_import
    return data


@router.post("/work-orders/{work_order_id}/import-links")
async def import_work_order_links(
    work_order_id: int,
    body: WorkOrderLinksBody,
    db: AsyncSession = Depends(get_db),
):
    """公司/取证：将文本链接或附件重新导入工单专属链接池"""
    wo = (await db.execute(select(WorkOrder).where(WorkOrder.id == work_order_id))).scalar_one_or_none()
    if not wo:
        raise HTTPException(404, "工单不存在")

    all_urls: list[str] = list(body.links or [])
    atts = (
        await db.execute(
            select(WorkOrderAttachment).where(
                WorkOrderAttachment.work_order_id == work_order_id,
                WorkOrderAttachment.file_type == "links",
            )
        )
    ).scalars().all()
    for att in atts:
        full = PROJECT_ROOT / att.file_path
        if full.is_file():
            all_urls.extend(_parse_links_from_bytes(full.read_bytes(), att.file_name))

    summary = await _import_links_to_batch(db, wo, all_urls)
    if wo.status == "submitted" and summary["imported"] > 0:
        wo.status = "collecting"
        if not wo.started_at:
            wo.started_at = datetime.now()
    await db.commit()
    return summary


@router.get("/work-orders/{work_order_id}/link-pool")
async def get_work_order_link_pool(work_order_id: int, db: AsyncSession = Depends(get_db)):
    wo = (await db.execute(select(WorkOrder).where(WorkOrder.id == work_order_id))).scalar_one_or_none()
    if not wo:
        raise HTTPException(404, "工单不存在")
    return await _batch_info_for_work_order(db, wo)
