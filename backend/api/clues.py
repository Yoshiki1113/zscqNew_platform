"""侵权线索黑名单 API"""
import os
from datetime import datetime

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy import delete, select, func
from sqlalchemy.ext.asyncio import AsyncSession

from database import get_db
from models import InfringementClue

router = APIRouter()


@router.get("/clues")
async def list_clues(
    keyword: str = "",
    page: int = 1,
    page_size: int = 50,
    db: AsyncSession = Depends(get_db),
):
    """查询侵权线索黑名单"""
    base = select(InfringementClue)
    count_base = select(func.count(InfringementClue.id))

    if keyword:
        like = f"%{keyword}%"
        cond = InfringementClue.account_name.like(like) | InfringementClue.work_name.like(like)
        base = base.where(cond)
        count_base = count_base.where(cond)

    total = (await db.execute(count_base)).scalar() or 0
    rows = (await db.execute(
        base.order_by(InfringementClue.id.desc())
        .offset((page - 1) * page_size).limit(page_size)
    )).scalars().all()

    return {
        "items": [
            {
                "id": r.id,
                "account_name": r.account_name or "",
                "work_name": r.work_name or "",
                "our_work_name": r.our_work_name or "",
                "company_name": r.company_name or "",
                "traffic_description": r.traffic_description or "",
                "video_link": r.video_link or "",
                "created_at": r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
            }
            for r in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("/clues/import")
async def import_clues_from_excel(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """从 Excel 文件导入侵权线索（会先清空旧数据）"""
    import openpyxl
    from io import BytesIO

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # 读取表头
    header_row = next(ws.iter_rows(values_only=True))
    col_map = {}
    for idx, val in enumerate(header_row):
        if val is None:
            continue
        name = str(val).strip()
        if "账号名称" in name:
            col_map["account_name"] = idx
        elif "侵权作品名称" in name:
            col_map["work_name"] = idx
        elif "我方剧名" in name or "我方作品名" in name:
            col_map["our_work_name"] = idx
        elif "账号主体公司" in name or "账号所属公司" in name or "公司" in name:
            col_map["company_name"] = idx
        elif "引流过程概述" in name or "引流" in name:
            col_map["traffic_description"] = idx
        elif "视频链接" in name or "链接" in name:
            col_map["video_link"] = idx

    if "account_name" not in col_map or "work_name" not in col_map:
        raise HTTPException(status_code=400, detail=f"Excel 缺少必要列，需要「账号名称」和「侵权作品名称」，实际表头: {header_row}")

    # 读取数据
    clues = []
    skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue

        account = (row[col_map["account_name"]] or "").strip() if len(row) > col_map["account_name"] else ""
        work = (row[col_map["work_name"]] or "").strip() if len(row) > col_map["work_name"] else ""

        if not account and not work:
            skipped += 1
            continue

        our_work = ""
        if "our_work_name" in col_map and len(row) > col_map["our_work_name"]:
            our_work = (row[col_map["our_work_name"]] or "").strip()
        company = ""
        if "company_name" in col_map and len(row) > col_map["company_name"]:
            company = (row[col_map["company_name"]] or "").strip()
        traffic_desc = ""
        if "traffic_description" in col_map and len(row) > col_map["traffic_description"]:
            traffic_desc = (row[col_map["traffic_description"]] or "").strip()
        link = ""
        if "video_link" in col_map and len(row) > col_map["video_link"]:
            link = (row[col_map["video_link"]] or "").strip()

        clues.append(InfringementClue(
            account_name=account,
            work_name=work,
            our_work_name=our_work,
            company_name=company,
            traffic_description=traffic_desc,
            video_link=link,
        ))

    wb.close()

    # 清空旧数据并批量插入
    await db.execute(delete(InfringementClue))
    if clues:
        db.add_all(clues)
    await db.commit()

    return {
        "imported": len(clues),
        "skipped": skipped,
        "db_total": len(clues),
        "imported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


@router.get("/clues/with-links")
async def get_clues_with_links(db: AsyncSession = Depends(get_db)):
    """查询有视频链接的线索数量和列表（用于前端显示"从线索创建任务"入口）"""
    result = await db.execute(
        select(InfringementClue).where(
            InfringementClue.video_link != "",
            InfringementClue.video_link.isnot(None),
        )
    )
    clues = result.scalars().all()
    total_result = await db.execute(select(func.count(InfringementClue.id)))
    return {
        "total_clues": total_result.scalar() or 0,
        "with_links": len(clues),
        "links": [
            {
                "id": c.id,
                "account_name": c.account_name or "",
                "work_name": c.work_name or "",
                "video_link": c.video_link or "",
            }
            for c in clues[:20]  # 只返回前 20 条预览
        ],
    }


@router.delete("/clues/all")
async def clear_clues(db: AsyncSession = Depends(get_db)):
    """清空所有侵权线索"""
    result = await db.execute(delete(InfringementClue))
    await db.commit()
    return {"deleted": result.rowcount or 0}
