"""侵权线索 Excel 导入脚本

用法:
    python import_clues.py [excel_path]

如果不传 excel_path，默认使用项目根目录下的「侵权线索.xlsx」。

功能:
    1. 读取 Excel 中的侵权线索
    2. 清空旧的 infringement_clues 表
    3. 批量导入到数据库
    4. 输出导入统计
"""
import asyncio
import os
import sys
from datetime import datetime

# 确保能导入 backend 模块
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from sqlalchemy import delete, select

from database import async_session, engine
from models import Base, InfringementClue


async def import_clues(excel_path: str) -> dict:
    """从 Excel 导入侵权线索到数据库。

    Returns:
        导入统计 dict
    """
    if not os.path.isfile(excel_path):
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    print(f"[导入] 读取 Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    # 读取表头，定位列索引
    header_row = next(ws.iter_rows(values_only=True))
    col_map = {}
    for idx, val in enumerate(header_row):
        if val is None:
            continue
        name = str(val).strip()
        if "账号名称" in name:
            col_map["account_name"] = idx
        elif "侵权作品名称" in name:
            col_map["work_name"] = idx
        elif "我方剧名" in name or "我方作品名" in name:
            col_map["our_work_name"] = idx
        elif "账号主体公司" in name or "账号所属公司" in name or "公司" in name:
            col_map["company_name"] = idx
        elif "引流过程概述" in name or "引流" in name:
            col_map["traffic_description"] = idx
        elif "视频链接" in name or "链接" in name:
            col_map["video_link"] = idx

    required = ["account_name", "work_name"]
    missing = [c for c in required if c not in col_map]
    if missing:
        raise ValueError(f"Excel 缺少必要列: {missing}，实际表头: {header_row}")

    print(f"[导入] 列映射: {col_map}")

    # 读取数据行
    clues = []
    skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:  # 跳过表头
            continue

        account = (row[col_map["account_name"]] or "").strip() if len(row) > col_map["account_name"] else ""
        work = (row[col_map["work_name"]] or "").strip() if len(row) > col_map["work_name"] else ""

        if not account and not work:
            skipped += 1
            continue

        our_work = ""
        if "our_work_name" in col_map and len(row) > col_map["our_work_name"]:
            our_work = (row[col_map["our_work_name"]] or "").strip()

        company = ""
        if "company_name" in col_map and len(row) > col_map["company_name"]:
            company = (row[col_map["company_name"]] or "").strip()

        traffic_desc = ""
        if "traffic_description" in col_map and len(row) > col_map["traffic_description"]:
            traffic_desc = (row[col_map["traffic_description"]] or "").strip()

        link = ""
        if "video_link" in col_map and len(row) > col_map["video_link"]:
            link = (row[col_map["video_link"]] or "").strip()

        clues.append(InfringementClue(
            account_name=account,
            work_name=work,
            our_work_name=our_work,
            company_name=company,
            traffic_description=traffic_desc,
            video_link=link,
        ))

    wb.close()

    print(f"[导入] 解析完成: {len(clues)} 条有效, {skipped} 条跳过(空行/重复)")

    # 写入数据库
    async with async_session() as session:
        # 先清空旧数据
        await session.execute(delete(InfringementClue))
        print("[导入] 已清空旧数据")

        # 批量插入
        if clues:
            session.add_all(clues)
            await session.commit()

        # 统计
        result = await session.execute(select(InfringementClue))
        total = len(result.scalars().all())

    stats = {
        "total_in_file": len(clues) + skipped,
        "imported": len(clues),
        "skipped": skipped,
        "db_total": total,
        "imported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    print(f"[导入] 完成: {stats}")
    return stats


async def main():
    # 确保表存在
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    excel_path = sys.argv[1] if len(sys.argv) > 1 else "../侵权线索.xlsx"
    if not os.path.isabs(excel_path):
        excel_path = os.path.join(os.path.dirname(__file__), excel_path)

    stats = await import_clues(excel_path)
    print(f"\n导入成功！数据库共有 {stats['db_total']} 条侵权线索")


if __name__ == "__main__":
    asyncio.run(main())
